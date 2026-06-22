"""
회원용 미수금 조회 사이트
강원도개인소형화물협회 / misugeum 확장 모듈

추가 라우트:
- GET  /member-arrears               회원 조회 화면
- POST /member-arrears               회원 조회 처리
- GET  /admin/public-arrears         관리자 공개자료 관리
- POST /admin/public-arrears/upload  공개조회용 엑셀/CSV 업로드
- POST /admin/public-arrears/sync    기존 DB에서 공개조회용 데이터 생성 시도

주의:
- 회원 화면에는 주소, 주민번호, 전체 전화번호를 절대 표시하지 않습니다.
- 조회 조건은 차량번호 + 성명 + 휴대폰 뒤4자리입니다.
- 관리자 업로드/동기화는 환경변수 PUBLIC_ARREARS_ADMIN_KEY가 있어야 작동합니다.
"""
from __future__ import annotations

import csv
import hashlib
import io
import os
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import MetaData, Table, and_, func, inspect, select, text
from sqlalchemy.engine import Engine
from sqlalchemy.orm import Session

try:
    from database import engine, get_db
except Exception as exc:  # pragma: no cover
    engine = None  # type: ignore
    get_db = None  # type: ignore
    _IMPORT_ERROR = exc
else:
    _IMPORT_ERROR = None

TEMPLATE_DIR = "templates" if Path("templates").exists() else "app/templates"
templates = Jinja2Templates(directory=TEMPLATE_DIR)
router = APIRouter()

# 메모리 기반 임시 차단. Railway 재시작 시 초기화됩니다.
_FAIL_BUCKET: Dict[str, Dict[str, Any]] = {}
MAX_FAILS = int(os.getenv("PUBLIC_ARREARS_MAX_FAILS", "5"))
BLOCK_SECONDS = int(os.getenv("PUBLIC_ARREARS_BLOCK_SECONDS", "900"))


# ─────────────────────────────────────────────────────────────
# 기본 유틸
# ─────────────────────────────────────────────────────────────
def _s(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _digits(v: Any) -> str:
    return re.sub(r"\D", "", _s(v))


def normalize_name(v: Any) -> str:
    return re.sub(r"\s+", "", _s(v))


def mask_name(name: str) -> str:
    name = normalize_name(name)
    if not name:
        return ""
    if len(name) <= 1:
        return name + "○"
    if len(name) == 2:
        return name[0] + "○"
    return name[0] + "○" * (len(name) - 2) + name[-1]


def normalize_vehicle(v: Any) -> str:
    """차량번호 정규화: 공백/하이픈/호 제거, 강원 접두 유지."""
    value = _s(v)
    value = value.replace(" ", "").replace("-", "").replace("ㆍ", "")
    value = value.replace("호", "")
    value = re.sub(r"[^0-9가-힣A-Za-z]", "", value)
    return value.lower()


def vehicle_last4(v: Any) -> str:
    d = _digits(v)
    return d[-4:] if len(d) >= 4 else d


def normalize_header(h: Any) -> str:
    return re.sub(r"\s+", "", _s(h)).lower()


def parse_amount(v: Any) -> int:
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = _s(v)
    if not s:
        return 0
    # 괄호 음수: (10,000) -> -10000
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
    if "-" in s[:2] or "▲" in s:
        neg = True
    n = re.sub(r"[^0-9]", "", s)
    if not n:
        return 0
    amount = int(n)
    return -amount if neg else amount


def comma(n: Any) -> str:
    try:
        return f"{int(n):,}"
    except Exception:
        return "0"


def hash_value(v: Any) -> str:
    salt = os.getenv("PUBLIC_ARREARS_HASH_SALT", "misugeum-public-arrears")
    return hashlib.sha256((salt + "|" + _s(v)).encode("utf-8")).hexdigest()


def client_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def now_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# 템플릿에서 쓸 필터
try:
    templates.env.filters["comma"] = comma
except Exception:
    pass


# ─────────────────────────────────────────────────────────────
# 테이블 생성
# ─────────────────────────────────────────────────────────────
def ensure_tables() -> None:
    if engine is None:
        raise RuntimeError(f"database.py를 불러오지 못했습니다: {_IMPORT_ERROR}")

    dialect = engine.dialect.name
    with engine.begin() as conn:
        if dialect == "postgresql":
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS public_arrears_lookup (
                    id SERIAL PRIMARY KEY,
                    vehicle_no VARCHAR(80) NOT NULL,
                    vehicle_norm VARCHAR(120) NOT NULL,
                    vehicle_last4 VARCHAR(10) NOT NULL,
                    name VARCHAR(80) NOT NULL,
                    name_norm VARCHAR(80) NOT NULL,
                    name_masked VARCHAR(80) NOT NULL,
                    phone_last4 VARCHAR(10) NOT NULL,
                    region VARCHAR(50),
                    account_type VARCHAR(30),
                    balance_total INTEGER DEFAULT 0,
                    association_fee_due INTEGER DEFAULT 0,
                    management_fee_due INTEGER DEFAULT 0,
                    age70_fee_due INTEGER DEFAULT 0,
                    prepaid_amount INTEGER DEFAULT 0,
                    base_month VARCHAR(20),
                    bank_account VARCHAR(120),
                    memo TEXT,
                    is_public BOOLEAN DEFAULT TRUE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS public_arrears_lookup_log (
                    id SERIAL PRIMARY KEY,
                    vehicle_input VARCHAR(120),
                    vehicle_last4 VARCHAR(10),
                    name_input_hash VARCHAR(128),
                    phone_last4_hash VARCHAR(128),
                    success BOOLEAN DEFAULT FALSE,
                    reason VARCHAR(80),
                    ip_hash VARCHAR(128),
                    user_agent TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))
        else:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS public_arrears_lookup (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    vehicle_no TEXT NOT NULL,
                    vehicle_norm TEXT NOT NULL,
                    vehicle_last4 TEXT NOT NULL,
                    name TEXT NOT NULL,
                    name_norm TEXT NOT NULL,
                    name_masked TEXT NOT NULL,
                    phone_last4 TEXT NOT NULL,
                    region TEXT,
                    account_type TEXT,
                    balance_total INTEGER DEFAULT 0,
                    association_fee_due INTEGER DEFAULT 0,
                    management_fee_due INTEGER DEFAULT 0,
                    age70_fee_due INTEGER DEFAULT 0,
                    prepaid_amount INTEGER DEFAULT 0,
                    base_month TEXT,
                    bank_account TEXT,
                    memo TEXT,
                    is_public INTEGER DEFAULT 1,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """))
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS public_arrears_lookup_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    vehicle_input TEXT,
                    vehicle_last4 TEXT,
                    name_input_hash TEXT,
                    phone_last4_hash TEXT,
                    success INTEGER DEFAULT 0,
                    reason TEXT,
                    ip_hash TEXT,
                    user_agent TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """))

        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_public_arrears_lookup_key ON public_arrears_lookup (vehicle_last4, name_norm, phone_last4)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_public_arrears_lookup_vehicle ON public_arrears_lookup (vehicle_norm)"))


try:
    if engine is not None:
        ensure_tables()
except Exception:
    # 앱 기동 자체가 깨지지 않게 라우트 접근 시 다시 시도합니다.
    pass


# ─────────────────────────────────────────────────────────────
# 보안 / 로그
# ─────────────────────────────────────────────────────────────
def check_blocked(request: Request, vehicle: str, name: str) -> Tuple[bool, int]:
    key = hash_value(client_ip(request) + "|" + vehicle_last4(vehicle) + "|" + normalize_name(name))
    row = _FAIL_BUCKET.get(key)
    if not row:
        return False, 0
    blocked_until = int(row.get("blocked_until") or 0)
    if blocked_until and blocked_until > int(time.time()):
        return True, blocked_until - int(time.time())
    return False, 0


def record_fail(request: Request, vehicle: str, name: str) -> None:
    key = hash_value(client_ip(request) + "|" + vehicle_last4(vehicle) + "|" + normalize_name(name))
    row = _FAIL_BUCKET.setdefault(key, {"count": 0, "blocked_until": 0})
    row["count"] += 1
    if row["count"] >= MAX_FAILS:
        row["blocked_until"] = int(time.time()) + BLOCK_SECONDS


def clear_fail(request: Request, vehicle: str, name: str) -> None:
    key = hash_value(client_ip(request) + "|" + vehicle_last4(vehicle) + "|" + normalize_name(name))
    _FAIL_BUCKET.pop(key, None)


def insert_lookup_log(
    db: Session,
    request: Request,
    vehicle_input: str,
    name_input: str,
    phone_last4_input: str,
    success: bool,
    reason: str,
) -> None:
    try:
        db.execute(text("""
            INSERT INTO public_arrears_lookup_log
            (vehicle_input, vehicle_last4, name_input_hash, phone_last4_hash, success, reason, ip_hash, user_agent, created_at)
            VALUES (:vehicle_input, :vehicle_last4, :name_hash, :phone_hash, :success, :reason, :ip_hash, :ua, :created_at)
        """), {
            "vehicle_input": normalize_vehicle(vehicle_input)[:120],
            "vehicle_last4": vehicle_last4(vehicle_input),
            "name_hash": hash_value(name_input),
            "phone_hash": hash_value(phone_last4_input),
            "success": bool(success),
            "reason": reason[:80],
            "ip_hash": hash_value(client_ip(request)),
            "ua": request.headers.get("user-agent", "")[:500],
            "created_at": now_ts(),
        })
        db.commit()
    except Exception:
        db.rollback()


def admin_key_ok(value: str) -> Tuple[bool, str]:
    expected = os.getenv("PUBLIC_ARREARS_ADMIN_KEY", "").strip()
    if not expected:
        return False, "환경변수 PUBLIC_ARREARS_ADMIN_KEY를 먼저 설정해야 관리자 기능을 사용할 수 있습니다."
    if not value or value.strip() != expected:
        return False, "관리자 키가 맞지 않습니다."
    return True, ""


# ─────────────────────────────────────────────────────────────
# 회원 조회 화면
# ─────────────────────────────────────────────────────────────
@router.get("/member-arrears", response_class=HTMLResponse)
def member_arrears_page(request: Request):
    return templates.TemplateResponse("member_arrears_lookup.html", {
        "request": request,
        "result": None,
        "error": None,
        "vehicle_no": "",
        "name": "",
        "phone_last4": "",
    })


@router.post("/member-arrears", response_class=HTMLResponse)
def member_arrears_lookup(
    request: Request,
    vehicle_no: str = Form(...),
    name: str = Form(...),
    phone_last4: str = Form(...),
    db: Session = Depends(get_db),
):
    ensure_tables()
    v_norm = normalize_vehicle(vehicle_no)
    v_last4 = vehicle_last4(vehicle_no)
    n_norm = normalize_name(name)
    p_last4 = _digits(phone_last4)[-4:]

    base_ctx = {
        "request": request,
        "result": None,
        "vehicle_no": vehicle_no,
        "name": name,
        "phone_last4": phone_last4,
    }

    if not v_last4 or len(v_last4) < 4 or not n_norm or len(p_last4) != 4:
        insert_lookup_log(db, request, vehicle_no, name, phone_last4, False, "invalid_input")
        return templates.TemplateResponse("member_arrears_lookup.html", {
            **base_ctx,
            "error": "차량번호, 성명, 휴대폰 뒤 4자리를 정확히 입력해 주세요.",
        })

    blocked, remain = check_blocked(request, vehicle_no, name)
    if blocked:
        insert_lookup_log(db, request, vehicle_no, name, phone_last4, False, "blocked")
        return templates.TemplateResponse("member_arrears_lookup.html", {
            **base_ctx,
            "error": f"조회 실패가 반복되어 잠시 제한되었습니다. 약 {max(1, remain // 60)}분 뒤 다시 시도해 주세요.",
        })

    rows = db.execute(text("""
        SELECT id, vehicle_no, vehicle_norm, vehicle_last4, name_masked, region, account_type,
               balance_total, association_fee_due, management_fee_due, age70_fee_due, prepaid_amount,
               base_month, bank_account, memo, updated_at
        FROM public_arrears_lookup
        WHERE is_public = TRUE
          AND vehicle_last4 = :vehicle_last4
          AND name_norm = :name_norm
          AND phone_last4 = :phone_last4
        ORDER BY updated_at DESC, id DESC
        LIMIT 10
    """), {
        "vehicle_last4": v_last4,
        "name_norm": n_norm,
        "phone_last4": p_last4,
    }).mappings().all()

    # 차량번호 전체 정규화가 일치하는 데이터가 있으면 그것만 우선 사용
    exact_rows = [r for r in rows if _s(r.get("vehicle_norm")) == v_norm]
    if exact_rows:
        rows = exact_rows

    if len(rows) == 1:
        clear_fail(request, vehicle_no, name)
        insert_lookup_log(db, request, vehicle_no, name, phone_last4, True, "success")
        result = dict(rows[0])
        return templates.TemplateResponse("member_arrears_lookup.html", {
            **base_ctx,
            "result": result,
            "error": None,
        })

    if len(rows) > 1:
        insert_lookup_log(db, request, vehicle_no, name, phone_last4, False, "multiple_match")
        return templates.TemplateResponse("member_arrears_lookup.html", {
            **base_ctx,
            "error": "일치하는 내역이 여러 건입니다. 정확한 확인은 협회로 문의해 주세요.",
        })

    record_fail(request, vehicle_no, name)
    insert_lookup_log(db, request, vehicle_no, name, phone_last4, False, "not_found")
    return templates.TemplateResponse("member_arrears_lookup.html", {
        **base_ctx,
        "error": "입력하신 정보와 일치하는 미수금 내역이 없습니다. 차량번호 또는 휴대폰번호가 변경된 경우 협회로 문의해 주세요.",
    })


# ─────────────────────────────────────────────────────────────
# 업로드 파싱
# ─────────────────────────────────────────────────────────────
def pick_value(row: Dict[str, Any], names: Iterable[str]) -> Any:
    norm_map = {normalize_header(k): v for k, v in row.items()}
    for name in names:
        key = normalize_header(name)
        if key in norm_map and _s(norm_map[key]):
            return norm_map[key]
    return ""


def find_last_arrears_amount(row: Dict[str, Any]) -> int:
    # 1월 미수금~12월 미수금처럼 월별 누적잔액 컬럼 중 마지막으로 값이 있는 컬럼을 현재잔액으로 사용
    candidates: List[Tuple[int, str, Any]] = []
    for i, (k, v) in enumerate(row.items()):
        hk = normalize_header(k)
        if not _s(v):
            continue
        if "미수금" in hk or "현재잔액" in hk or "미수잔액" in hk or hk in {"미수", "금액", "잔액", "balance", "amount"}:
            candidates.append((i, hk, v))
    if not candidates:
        return 0

    # 월별 미수금이 있으면 컬럼 순서상 가장 뒤의 값을 현재잔액으로 본다.
    candidates.sort(key=lambda x: x[0])
    return parse_amount(candidates[-1][2])


def normalize_account_type(v: Any, vehicle: str = "") -> str:
    s = _s(v)
    if "협회가입" in s:
        return "협회비"
    if "택배신규" in s:
        return "관리비"
    if "협회" in s:
        return "협회비"
    if "관리" in s:
        return "관리비"
    if "70" in s:
        return "70세"
    if "배" in _s(vehicle):
        return "관리비"
    return s or "미수금"


def rows_from_csv(content: bytes) -> List[Dict[str, Any]]:
    for enc in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            text_content = content.decode(enc)
            break
        except Exception:
            continue
    else:
        text_content = content.decode("utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(text_content))
    return [dict(r) for r in reader]


def rows_from_excel(content: bytes, filename: str) -> List[Dict[str, Any]]:
    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError("openpyxl이 설치되어 있어야 엑셀 업로드가 가능합니다.") from exc

    keep_vba = filename.lower().endswith(".xlsm")
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True, keep_vba=keep_vba)
    sheet_name = None
    for wanted in ("회원조회용", "2026년회비내역", "미수금", "Sheet1"):
        if wanted in wb.sheetnames:
            sheet_name = wanted
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    values = list(ws.iter_rows(values_only=True))
    header_idx = 0
    for idx, row in enumerate(values[:30]):
        joined = " ".join(_s(c) for c in row)
        if ("차량" in joined or "차번" in joined) and ("성명" in joined or "이름" in joined or "회원" in joined):
            header_idx = idx
            break

    headers = [_s(h) or f"col_{i}" for i, h in enumerate(values[header_idx])]
    rows: List[Dict[str, Any]] = []
    for raw in values[header_idx + 1:]:
        if not any(_s(v) for v in raw):
            continue
        row = {headers[i]: raw[i] if i < len(raw) else "" for i in range(len(headers))}
        rows.append(row)
    return rows


def convert_public_rows(raw_rows: List[Dict[str, Any]], base_month_default: str = "") -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    out: List[Dict[str, Any]] = []
    stats = {"total": 0, "converted": 0, "skipped": 0, "missing_phone": 0, "zero_or_negative": 0}

    for row in raw_rows:
        stats["total"] += 1
        vehicle = pick_value(row, ["차량번호", "차량 번호", "차번", "등록번호", "vehicle_no", "vehicle"])
        name = pick_value(row, ["성명", "이름", "회원명", "대표자", "차주명", "name"])
        phone = pick_value(row, ["핸드폰", "휴대폰", "휴대전화", "전화번호", "연락처", "phone", "mobile"])
        region = pick_value(row, ["지역", "시군", "region"])
        account_type_raw = pick_value(row, ["계정", "구분", "부과구분", "회원구분", "account_type"])
        base_month = pick_value(row, ["기준월", "기준 월", "yyyymm", "base_month"]) or base_month_default

        if not vehicle or not name:
            stats["skipped"] += 1
            continue

        amount = find_last_arrears_amount(row)
        account_type = normalize_account_type(account_type_raw, _s(vehicle))
        phone_last4_value = _digits(phone)[-4:]
        if len(phone_last4_value) != 4:
            stats["missing_phone"] += 1
            # 보안상 휴대폰 뒤4자리 없는 행은 공개 조회용으로 넣지 않습니다.
            stats["skipped"] += 1
            continue

        if amount <= 0:
            stats["zero_or_negative"] += 1

        association_due = amount if account_type == "협회비" and amount > 0 else 0
        management_due = amount if account_type == "관리비" and amount > 0 else 0
        age70_due = amount if "70" in account_type and amount > 0 else 0
        prepaid = abs(amount) if amount < 0 else 0

        out.append({
            "vehicle_no": _s(vehicle),
            "vehicle_norm": normalize_vehicle(vehicle),
            "vehicle_last4": vehicle_last4(vehicle),
            "name": normalize_name(name),
            "name_norm": normalize_name(name),
            "name_masked": mask_name(_s(name)),
            "phone_last4": phone_last4_value,
            "region": _s(region),
            "account_type": account_type,
            "balance_total": amount,
            "association_fee_due": association_due,
            "management_fee_due": management_due,
            "age70_fee_due": age70_due,
            "prepaid_amount": prepaid,
            "base_month": _s(base_month),
            "bank_account": os.getenv("PUBLIC_ARREARS_BANK_ACCOUNT", "농협 계좌번호를 관리자 화면에서 설정해 주세요"),
            "memo": "",
        })
        stats["converted"] += 1

    return out, stats


def replace_public_rows(db: Session, rows: List[Dict[str, Any]], mode: str = "replace") -> int:
    ensure_tables()
    if mode == "replace":
        db.execute(text("DELETE FROM public_arrears_lookup"))

    insert_sql = text("""
        INSERT INTO public_arrears_lookup
        (vehicle_no, vehicle_norm, vehicle_last4, name, name_norm, name_masked, phone_last4,
         region, account_type, balance_total, association_fee_due, management_fee_due, age70_fee_due,
         prepaid_amount, base_month, bank_account, memo, is_public, created_at, updated_at)
        VALUES
        (:vehicle_no, :vehicle_norm, :vehicle_last4, :name, :name_norm, :name_masked, :phone_last4,
         :region, :account_type, :balance_total, :association_fee_due, :management_fee_due, :age70_fee_due,
         :prepaid_amount, :base_month, :bank_account, :memo, TRUE, :created_at, :updated_at)
    """)
    now = now_ts()
    count = 0
    for r in rows:
        data = dict(r)
        data["created_at"] = now
        data["updated_at"] = now
        db.execute(insert_sql, data)
        count += 1
    db.commit()
    return count


# ─────────────────────────────────────────────────────────────
# 기존 DB에서 공개자료 생성 시도
# ─────────────────────────────────────────────────────────────
def _choose_col(cols: Iterable[str], names: Iterable[str]) -> Optional[str]:
    col_list = list(cols)
    norm_to_real = {normalize_header(c): c for c in col_list}
    for name in names:
        key = normalize_header(name)
        if key in norm_to_real:
            return norm_to_real[key]
    for c in col_list:
        nc = normalize_header(c)
        for name in names:
            if normalize_header(name) in nc:
                return c
    return None


def sync_public_rows_from_existing_db(db: Session) -> Tuple[int, str]:
    """프로젝트마다 테이블명이 달라도 최대한 찾아서 공개조회용 데이터를 생성합니다."""
    if engine is None:
        return 0, "DB 엔진을 불러오지 못했습니다."
    ensure_tables()
    inspector = inspect(engine)
    tables = inspector.get_table_names()
    metadata = MetaData()

    member_candidates = []
    arrears_candidates = []
    for table_name in tables:
        try:
            columns = [c["name"] for c in inspector.get_columns(table_name)]
        except Exception:
            continue
        v_col = _choose_col(columns, ["vehicle_no", "vehicle_number", "차량번호", "차번", "등록번호"])
        n_col = _choose_col(columns, ["name", "성명", "회원명", "대표자"])
        p_col = _choose_col(columns, ["phone", "mobile", "핸드폰", "휴대폰", "전화번호", "연락처"])
        a_col = _choose_col(columns, ["balance", "current_balance", "amount", "arrears", "excel_arrears", "미수금", "현재잔액", "미수잔액"])
        if v_col and n_col and p_col:
            member_candidates.append((table_name, columns, v_col, n_col, p_col))
        if a_col:
            arrears_candidates.append((table_name, columns, a_col))

    # 1) 한 테이블에 차량/성명/전화/금액이 모두 있으면 그 테이블 우선
    for table_name, columns, amount_col in arrears_candidates:
        v_col = _choose_col(columns, ["vehicle_no", "vehicle_number", "차량번호", "차번", "등록번호"])
        n_col = _choose_col(columns, ["name", "성명", "회원명", "대표자"])
        p_col = _choose_col(columns, ["phone", "mobile", "핸드폰", "휴대폰", "전화번호", "연락처"])
        if v_col and n_col and p_col:
            t = Table(table_name, metadata, autoload_with=engine)
            region_col = _choose_col(columns, ["region", "지역", "시군"])
            account_col = _choose_col(columns, ["account_type", "account", "계정", "구분", "부과구분"])
            base_col = _choose_col(columns, ["base_month", "yyyymm", "기준월"])
            stmt = select(
                t.c[v_col].label("vehicle_no"),
                t.c[n_col].label("name"),
                t.c[p_col].label("phone"),
                t.c[amount_col].label("amount"),
                (t.c[region_col].label("region") if region_col else text("'' AS region")),
                (t.c[account_col].label("account_type") if account_col else text("'' AS account_type")),
                (t.c[base_col].label("base_month") if base_col else text("'' AS base_month")),
            ).limit(20000)
            raw = [dict(r._mapping) for r in db.execute(stmt).all()]
            public_rows, stats = convert_public_rows([
                {"차량번호": r.get("vehicle_no"), "성명": r.get("name"), "핸드폰": r.get("phone"),
                 "미수금": r.get("amount"), "지역": r.get("region"), "계정": r.get("account_type"), "기준월": r.get("base_month")}
                for r in raw
            ])
            saved = replace_public_rows(db, public_rows, mode="replace")
            return saved, f"{table_name} 테이블에서 공개조회용 데이터 {saved}건 생성. 전화번호 누락 제외 {stats.get('missing_phone', 0)}건."

    # 2) 회원 테이블 + 미수금 테이블 조인 시도
    for member_table_name, member_cols, v_col, n_col, p_col in member_candidates:
        m = Table(member_table_name, metadata, autoload_with=engine)
        m_id_col = _choose_col(member_cols, ["id", "member_id"])
        if not m_id_col:
            continue
        for arrears_table_name, arrears_cols, amount_col in arrears_candidates:
            if arrears_table_name == member_table_name:
                continue
            a_member_col = _choose_col(arrears_cols, ["member_id", "memberid", "member"])
            if not a_member_col:
                continue
            a = Table(arrears_table_name, metadata, autoload_with=engine)
            region_col = _choose_col(member_cols, ["region", "지역", "시군"])
            account_col = _choose_col(arrears_cols, ["account_type", "account", "계정", "구분", "부과구분"])
            base_col = _choose_col(arrears_cols, ["base_month", "yyyymm", "기준월", "year_month"])
            stmt = select(
                m.c[v_col].label("vehicle_no"),
                m.c[n_col].label("name"),
                m.c[p_col].label("phone"),
                a.c[amount_col].label("amount"),
                (m.c[region_col].label("region") if region_col else text("'' AS region")),
                (a.c[account_col].label("account_type") if account_col else text("'' AS account_type")),
                (a.c[base_col].label("base_month") if base_col else text("'' AS base_month")),
            ).select_from(m.join(a, m.c[m_id_col] == a.c[a_member_col])).limit(20000)
            raw = [dict(r._mapping) for r in db.execute(stmt).all()]
            if not raw:
                continue
            public_rows, stats = convert_public_rows([
                {"차량번호": r.get("vehicle_no"), "성명": r.get("name"), "핸드폰": r.get("phone"),
                 "미수금": r.get("amount"), "지역": r.get("region"), "계정": r.get("account_type"), "기준월": r.get("base_month")}
                for r in raw
            ])
            saved = replace_public_rows(db, public_rows, mode="replace")
            return saved, f"{member_table_name}+{arrears_table_name} 조인으로 공개조회용 데이터 {saved}건 생성. 전화번호 누락 제외 {stats.get('missing_phone', 0)}건."

    return 0, "기존 DB에서 차량번호+성명+휴대폰+미수금 구조를 자동으로 찾지 못했습니다. 관리자 화면에서 회원조회용 엑셀/CSV를 업로드해 주세요."


# ─────────────────────────────────────────────────────────────
# 관리자 화면
# ─────────────────────────────────────────────────────────────
@router.get("/admin/public-arrears", response_class=HTMLResponse)
def public_arrears_admin(request: Request, db: Session = Depends(get_db)):
    ensure_tables()
    total = db.execute(text("SELECT COUNT(*) FROM public_arrears_lookup")).scalar() or 0
    recent_logs = db.execute(text("""
        SELECT created_at, vehicle_last4, success, reason
        FROM public_arrears_lookup_log
        ORDER BY id DESC
        LIMIT 20
    """)).mappings().all()
    return templates.TemplateResponse("admin_public_arrears.html", {
        "request": request,
        "total": total,
        "recent_logs": recent_logs,
        "message": None,
        "error": None,
        "admin_key_set": bool(os.getenv("PUBLIC_ARREARS_ADMIN_KEY", "").strip()),
    })


@router.post("/admin/public-arrears/upload", response_class=HTMLResponse)
async def public_arrears_upload(
    request: Request,
    admin_key: str = Form(...),
    mode: str = Form("replace"),
    base_month: str = Form(""),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    ok, err = admin_key_ok(admin_key)
    if not ok:
        return templates.TemplateResponse("admin_public_arrears.html", {
            "request": request, "total": 0, "recent_logs": [], "message": None, "error": err,
            "admin_key_set": bool(os.getenv("PUBLIC_ARREARS_ADMIN_KEY", "").strip()),
        })

    content = await file.read()
    filename = file.filename or ""
    try:
        if filename.lower().endswith(".csv"):
            raw_rows = rows_from_csv(content)
        else:
            raw_rows = rows_from_excel(content, filename)
        public_rows, stats = convert_public_rows(raw_rows, base_month_default=base_month)
        saved = replace_public_rows(db, public_rows, mode=mode)
        total = db.execute(text("SELECT COUNT(*) FROM public_arrears_lookup")).scalar() or 0
        msg = f"업로드 완료: 저장 {saved}건 / 원본 {stats['total']}행 / 제외 {stats['skipped']}행 / 휴대폰 뒤4자리 없음 {stats['missing_phone']}행"
        return templates.TemplateResponse("admin_public_arrears.html", {
            "request": request, "total": total, "recent_logs": [], "message": msg, "error": None,
            "admin_key_set": True,
        })
    except Exception as exc:
        db.rollback()
        total = db.execute(text("SELECT COUNT(*) FROM public_arrears_lookup")).scalar() or 0
        return templates.TemplateResponse("admin_public_arrears.html", {
            "request": request, "total": total, "recent_logs": [], "message": None, "error": f"업로드 실패: {exc}",
            "admin_key_set": True,
        })


@router.post("/admin/public-arrears/sync", response_class=HTMLResponse)
def public_arrears_sync(
    request: Request,
    admin_key: str = Form(...),
    db: Session = Depends(get_db),
):
    ok, err = admin_key_ok(admin_key)
    if not ok:
        return templates.TemplateResponse("admin_public_arrears.html", {
            "request": request, "total": 0, "recent_logs": [], "message": None, "error": err,
            "admin_key_set": bool(os.getenv("PUBLIC_ARREARS_ADMIN_KEY", "").strip()),
        })
    try:
        saved, message = sync_public_rows_from_existing_db(db)
        total = db.execute(text("SELECT COUNT(*) FROM public_arrears_lookup")).scalar() or 0
        return templates.TemplateResponse("admin_public_arrears.html", {
            "request": request, "total": total, "recent_logs": [], "message": message, "error": None if saved else message,
            "admin_key_set": True,
        })
    except Exception as exc:
        db.rollback()
        total = db.execute(text("SELECT COUNT(*) FROM public_arrears_lookup")).scalar() or 0
        return templates.TemplateResponse("admin_public_arrears.html", {
            "request": request, "total": total, "recent_logs": [], "message": None, "error": f"DB 동기화 실패: {exc}",
            "admin_key_set": True,
        })

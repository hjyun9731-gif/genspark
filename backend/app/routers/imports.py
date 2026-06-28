"""엑셀 업로드/DB 반영 라우터.

핵심 업무 규칙
- 전체면허자현황은 반드시 '개인' + '택배' 시트만 읽는다.
- 업체세분/선진물류/업체/차량집계는 전체자명단 업로드에서 절대 사용하지 않는다.
- 2026미수금은 반드시 '2026년회비내역' 시트만 읽는다.
- 2026미수금은 회원 원장이 아니며, 이미 저장된 회원에게 현재 미수 잔액만 붙이는 보조 파일이다.
- 월별 미수금은 누적 잔액이므로 합산하지 않고, 행별 마지막 입력 월의 미수금 1건만 저장한다.
- 기존 데이터는 일반 업로드에서 삭제하지 않는다. 초기화 버튼을 눌렀을 때만 misu_* 업무 테이블을 비운다.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any

import pandas as pd
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from ..billing import charge_item, monthly_charge, next_month_ym
from ..database import get_db
from ..models import Closure, Deposit, Member, MemberHistory, Payment, Pending, ReceivableItem

router = APIRouter(prefix="/api/import", tags=["import"])

SIGUN = [
    "춘천시", "원주시", "강릉시", "동해시", "태백시", "속초시", "삼척시",
    "홍천군", "횡성군", "영월군", "평창군", "정선군", "철원군", "화천군",
    "양구군", "인제군", "고성군", "양양군",
]

STANDARD_MEMBER_COLUMNS = [
    "지역", "회원구분", "관리번호", "차량번호", "성명", "주민등록번호", "주소", "전화번호", "핸드폰",
    "인가일자", "가입일자", "자격증명 발급일자", "자격증명 발급번호", "운전면허증번호",
    "차종", "유종", "사업자등록번호", "소속업체", "공문 주소", "대리인", "구조변경",
    "비고", "전화 메모", "비고2", "비고3", "가입여부", "부과구분", "부과금액", "부과시작월",
]

ARREARS_PREVIEW_COLUMNS = [
    "지역", "계정", "비고", "차량번호", "성명", "대수", "이월금",
    "마지막 미수 기준월", "현재 미수금액", "매칭상태", "매칭된 회원명", "매칭된 차량번호",
]

DEPOSIT_PREVIEW_COLUMNS = [
    "거래일자", "입금자명", "입금액", "거래내용", "거래기록사항", "원본구분", "원본메모",
]

FORBIDDEN_MEMBER_SHEETS = {"업체세분", "선진물류", "업체", "차량집계"}


def _clean(v: Any) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).replace("\u3000", " ").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.strip()


def _norm_col(v: Any) -> str:
    return re.sub(r"\s+", "", _clean(v).replace("\n", ""))


def _person_name(v: Any) -> str:
    return re.sub(r"\s+", "", _clean(v))


def _clip(v: Any, n: int) -> str | None:
    s = _clean(v)
    return s[:n] if s else None


def _money(v: Any) -> int:
    s = _clean(v)
    if not s or s in {"-", "–", "—", "※"}:
        return 0
    s = re.sub(r"[^0-9\-]", "", s)
    try:
        return max(0, int(s or 0))
    except Exception:
        return 0


def _has_value(v: Any) -> bool:
    s = _clean(v)
    return bool(s) and s not in {"-", "–", "—"}


def _append_unique_note(existing: str | None, label: str, value: Any, max_len: int = 1800) -> str | None:
    note = _clean(value)
    if not note:
        return existing
    part = f"{label}:{note}"
    current = _clean(existing)
    # 같은 미수금 비고를 여러 번 업로드해도 계속 중복 추가하지 않는다.
    if part in current:
        return current[:max_len] if current else None
    combined = f"{current} / {part}" if current else part
    return combined[:max_len] if combined else None


def _strip_generated_member_memo(existing: str | None) -> str | None:
    """회원상세 표시용 기본정보(주소/공문주소/주민번호 등)를 메모에 보존한다.
    기존 버전처럼 주소 메모를 지우면 회원상세에서 주소가 '-'로 표시되므로 삭제하지 않는다.
    """
    current = _clean(existing)
    return current[:1800] if current else None


def _member_basic_note(row: dict[str, Any]) -> str | None:
    parts: list[str] = []
    # DB 모델에 별도 주소 컬럼이 없어도 회원상세/API에서 복원할 수 있게 메모에 구조화 저장한다.
    for k in ["주소", "공문 주소", "주민등록번호", "전화번호", "핸드폰", "자격증명 발급번호"]:
        v = _clean(row.get(k))
        if v:
            parts.append(f"{k}:{v}")
    return " / ".join(parts)[:1200] if parts else None


def _ledger_note(row: dict[str, Any]) -> str | None:
    """회원원장 업로드에서는 주소 같은 기본정보를 메모에 넣지 않는다.
    실제 비고류 칸만 따로 보존한다.
    """
    parts: list[str] = []
    basic_note = _member_basic_note(row)
    if basic_note:
        parts.append(basic_note)
    for k in ["비고", "비고2", "비고3", "전화 메모"]:
        v = _clean(row.get(k))
        if v:
            parts.append(f"원장 {k}:{v}")
    return " / ".join(parts)[:1800] if parts else None




def _merge_member_notes(existing: str | None, generated: str | None, max_len: int = 1800) -> str | None:
    """주소/공문주소/주민번호 등 구조화 메모를 원장 비고로 감싸지 않고 그대로 보존한다."""
    current = _clean(existing)
    note = _clean(generated)
    if not note:
        return current[:max_len] if current else None
    parts = [x.strip() for x in (current + " / " + note if current else note).split(" / ") if x.strip()]
    merged: list[str] = []
    seen: set[str] = set()
    for part in parts:
        if part in seen:
            continue
        seen.add(part)
        merged.append(part)
    return " / ".join(merged)[:max_len] if merged else None

def _is_contact_problem_note(value: Any) -> bool:
    t = _clean(value).replace(" ", "").lower()
    if not t:
        return False
    return any(key in t for key in ["결번", "반송", "연락두절", "전화안됨", "주소불명", "문자발송x", "지로x,반송"])


def _fix_reversed_future_date(d: date | None) -> date | None:
    """엑셀의 20.10.30 / 21.03.31 같은 날짜가 2030-10-20 / 2031-03-21로 뒤집힌 경우 보정."""
    if not d:
        return None
    yy = d.year % 100
    # 2030-10-20 -> 2020-10-30, 2031-03-21 -> 2021-03-31
    if d.year > date.today().year and 1 <= yy <= 31 and 1 <= d.day <= 31:
        try:
            fixed = date(2000 + d.day, d.month, yy)
            if fixed.year <= date.today().year:
                return fixed
        except Exception:
            pass
    return d


def _parse_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return _fix_reversed_future_date(v.date())
    if isinstance(v, date):
        return _fix_reversed_future_date(v)
    s = _clean(v)
    if not s or s.lower() == "x":
        return None

    # 통장/카드 엑셀에 26.03.31, 25.12.30 같은 형식이 많다.
    # pandas가 이를 2031-03-26처럼 잘못 해석하는 경우가 있어 직접 먼저 처리한다.
    # 기본은 YY.MM.DD로 보고, 해석한 연도가 너무 미래이면 DD.MM.YY로 보정한다.
    m = re.fullmatch(r"\s*(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{1,2})\s*", s)
    if m:
        a, b, c = map(int, m.groups())
        current_year = date.today().year
        try:
            yy_year = 2000 + a if a < 80 else 1900 + a
            parsed = date(yy_year, b, c)
            if parsed.year <= current_year + 1:
                return parsed
        except Exception:
            pass
        try:
            dd_year = 2000 + c if c < 80 else 1900 + c
            parsed = date(dd_year, b, a)
            return parsed
        except Exception:
            pass

    digits = re.sub(r"\D", "", s)
    try:
        if len(digits) == 8:
            return date(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))
        if len(digits) == 6:
            yy = int(digits[:2])
            year = 2000 + yy if yy < 80 else 1900 + yy
            parsed = date(year, int(digits[2:4]), int(digits[4:6]))
            if parsed.year <= date.today().year + 1:
                return parsed
            # 300326처럼 들어온 경우: 30.03.26 = 2026-03-30으로 보정
            return date(2000 + int(digits[4:6]), int(digits[2:4]), int(digits[:2]))
    except Exception:
        return None

    try:
        d = pd.to_datetime(s, errors="coerce")
        if pd.notna(d):
            return _fix_reversed_future_date(d.date())
    except Exception:
        pass
    return None


def _json(v: Any) -> Any:
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    return _clean(v)


def _sigun_from_text(text: str) -> str:
    t = _clean(text).replace(" ", "")
    for s in SIGUN:
        if s.replace(" ", "") in t or s[:-1] in t:
            return s
    return "미분류"


def _vehicle_norm(vehicle: str) -> str:
    s = _clean(vehicle)
    s = s.replace("호", "")
    s = re.sub(r"[\s\-_/.,()\[\]]+", "", s)
    return s


def _vehicle_last4(vehicle: str) -> str:
    nums = re.findall(r"\d+", _clean(vehicle))
    if not nums:
        return ""
    return nums[-1][-4:]


def _valid_vehicle(vehicle: str) -> bool:
    v = _clean(vehicle)
    if not v or "?" in v:
        return False
    last4 = _vehicle_last4(v)
    return len(last4) >= 4 and not re.fullmatch(r"0{4,}", last4)


def _birth_year(rrn: str) -> int | None:
    s = re.sub(r"\D", "", _clean(rrn))
    if len(s) < 7:
        return None
    yy = int(s[:2])
    marker = s[6]
    if marker in {"1", "2", "5", "6"}:
        return 1900 + yy
    if marker in {"3", "4", "7", "8"}:
        return 2000 + yy
    return 1900 + yy if yy > 30 else 2000 + yy


def _membership_from(row: dict[str, Any]) -> str:
    val = _clean(row.get("가입여부"))
    join = _clean(row.get("가입일자"))
    negative = {"x", "X", "×", "미가입", "비가입", "없음", "-"}
    if val in negative or "미가입" in val:
        return "협회미가입"
    if "가입" in val and "미가입" not in val:
        return "협회가입"
    if join and join not in negative and "미가입" not in join:
        return "협회가입"
    return "협회미가입"


def _next_member_id_from_no(no: int) -> str:
    return f"M{no:05d}"


def _current_max_member_no(db: Session) -> int:
    ids = db.scalars(select(Member.id).where(Member.id.like("M%"))).all()
    max_no = 0
    for mid in ids:
        m = re.search(r"(\d+)$", str(mid or ""))
        if m:
            max_no = max(max_no, int(m.group(1)))
    return max_no


def _make_mgmt_no(raw: str) -> str | None:
    s = _clean(raw)
    return s[:16] if s else None


def _read_workbook(file_bytes: bytes):
    return load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True, keep_links=False)


def _header_map(ws, max_col: int = 80, max_scan_rows: int = 12) -> tuple[int, list[str], dict[str, int]]:
    """헤더가 1행이 아닌 파일도 대비해 상단 여러 줄을 훑는다."""
    best_row_no = 1
    best_headers: list[str] = []
    best_mapping: dict[str, int] = {}
    best_score = -1
    expected = {"지역", "관리번호", "차량번호", "성명", "주민등록번호", "주소", "핸드폰", "가입일자", "자격증명발급일자"}
    for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, max_col=max_col, values_only=True), start=1):
        headers: list[str] = []
        mapping: dict[str, int] = {}
        for idx, v in enumerate(row, start=1):
            h = _clean(v)
            if not h:
                continue
            h = re.sub(r"\s+", " ", h.replace("\n", " ")).strip()
            headers.append(h)
            mapping[_norm_col(h)] = idx
        score = sum(1 for key in expected if key in mapping)
        if score > best_score:
            best_score = score
            best_row_no, best_headers, best_mapping = row_no, headers, mapping
        if score >= 4 and "차량번호" in mapping:
            break
    return best_row_no, best_headers, best_mapping


def _get(row: tuple, mapping: dict[str, int], *keys: str) -> Any:
    for key in keys:
        idx = mapping.get(_norm_col(key))
        if idx is not None and idx - 1 < len(row):
            return row[idx - 1]
    return None


def _iter_license_rows(file_bytes: bytes, preview_limit: int | None = None) -> list[dict[str, Any]]:
    wb = _read_workbook(file_bytes)
    rows: list[dict[str, Any]] = []
    missing = [s for s in ("개인", "택배") if s not in wb.sheetnames]
    if missing:
        raise HTTPException(status_code=400, detail=f"전체면허자현황에서 {', '.join(missing)} 시트를 찾지 못했습니다.")

    for sheet_name, member_type in (("개인", "개인"), ("택배", "택배")):
        ws = wb[sheet_name]
        header_row_no, _, m = _header_map(ws, max_col=80)
        # 안전장치: 업체세분 컬럼이면 즉시 실패시켜 잘못된 시트를 못 쓰게 한다.
        if "업체명및대표자" in m or "기사성명" in m or "입사일자" in m:
            raise HTTPException(status_code=400, detail=f"{sheet_name} 시트가 아니라 업체세분 형식 컬럼이 감지되었습니다. 개인/택배 시트만 업로드해야 합니다.")
        for row in ws.iter_rows(min_row=header_row_no + 1, max_col=80, values_only=True):
            vehicle = _clean(_get(row, m, "차량번호"))
            name = _person_name(_get(row, m, "성명", "성 명", "성    명"))
            if not _valid_vehicle(vehicle) or not name:
                continue
            region = _clean(_get(row, m, "지역", "지 역")) or _sigun_from_text(_clean(_get(row, m, "주소", "주 소")))
            if region.replace(" ", "") in {"합계", "총계"}:
                continue
            join_raw = _get(row, m, "가입일자")
            cert_raw = _get(row, m, "자격증명 발급일자", "자격증명 발급일", "자격증명\n발급일자")
            membership = _membership_from({"가입여부": _get(row, m, "가입여부"), "가입일자": join_raw})
            cert_date = _parse_date(cert_raw)
            join_date = _parse_date(join_raw)
            item = charge_item(membership)
            byear = _birth_year(_clean(_get(row, m, "주민등록번호")))
            age = (date.today().year - byear) if byear else None
            amount = monthly_charge(membership, age=age, birth_year=byear)
            row_out = {
                "지역": region or "미분류",
                "회원구분": member_type,
                "관리번호": _clean(_get(row, m, "관리번호")),
                "차량번호": vehicle,
                "성명": name,
                "주민등록번호": _clean(_get(row, m, "주민등록번호")),
                "주소": _clean(_get(row, m, "주소", "주 소")),
                "전화번호": _clean(_get(row, m, "전화번호")),
                "핸드폰": _clean(_get(row, m, "핸드폰", "핸 드 폰")),
                "인가일자": _json(_get(row, m, "인가일자")),
                "가입일자": _json(join_raw),
                "자격증명 발급일자": _json(cert_raw),
                "자격증명 발급번호": _clean(_get(row, m, "자격증명 발급번호", "자격증명\n발급번호")),
                "운전면허증번호": _clean(_get(row, m, "운전면허증번호")),
                "차종": _clean(_get(row, m, "차종")),
                "유종": _clean(_get(row, m, "유종")),
                "사업자등록번호": _clean(_get(row, m, "사업자등록번호")),
                "소속업체": _clean(_get(row, m, "소속업체")),
                "공문 주소": _clean(_get(row, m, "공문 주소", "공문주소")),
                "대리인": _clean(_get(row, m, "대리인")),
                "구조변경": _clean(_get(row, m, "구조변경")),
                "비고": _clean(_get(row, m, "비고")),
                "전화 메모": _clean(_get(row, m, "전화 메모", "전화메모")),
                "비고2": _clean(_get(row, m, "비고2", "비고2 ")),
                "비고3": _clean(_get(row, m, "비고3")),
                "가입여부": membership,
                "부과구분": item,
                "부과금액": amount,
                "부과시작월": next_month_ym(join_date if membership == "협회가입" else cert_date),
            }
            rows.append(row_out)
            if preview_limit and len(rows) >= preview_limit:
                return rows
    return rows


def _arrears_header(ws) -> tuple[int, list[str], dict[str, int]]:
    best = (1, [], {})
    for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=10, max_col=80, values_only=True), start=1):
        mapping = {}
        headers = []
        for i, v in enumerate(row, start=1):
            h = _clean(v)
            if not h:
                continue
            h = re.sub(r"\s+", " ", h.replace("\n", " ")).strip()
            headers.append(h)
            # 같은 제목(특히 비고)이 여러 번 있을 수 있다.
            # 미수금 파일의 입금자 별칭 비고는 앞쪽(계정 다음, 차량번호 앞) 비고이므로
            # 뒤쪽 중복 비고가 앞쪽 비고를 덮어쓰지 않게 첫 번째 위치를 유지한다.
            mapping.setdefault(_norm_col(h), i - 1)
        if "차량번호" in mapping and ("성명" in mapping or "성명" in [_norm_col(x) for x in headers]):
            return row_no, headers, mapping
        score = sum(1 for key in ["지역", "계정", "차량번호", "성명", "이월금"] if key in mapping)
        if score > sum(1 for key in ["지역", "계정", "차량번호", "성명", "이월금"] if key in best[2]):
            best = (row_no, headers, mapping)
    return best


def _iter_arrears_rows(file_bytes: bytes, preview_limit: int | None = None) -> list[dict[str, Any]]:
    wb = _read_workbook(file_bytes)
    if "2026년회비내역" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="미수금 파일에서 '2026년회비내역' 시트를 찾지 못했습니다. 2026년/월별부과대수/2025년(2) 시트는 사용하지 않습니다.")
    ws = wb["2026년회비내역"]
    header_row_no, headers, m = _arrears_header(ws)

    def idx(*labels: str) -> int | None:
        for label in labels:
            found = m.get(_norm_col(label))
            if found is not None:
                return found
        return None

    base = {
        "지역": idx("지역"), "계정": idx("계정"), "비고": idx("비고"),
        "차량번호": idx("차량번호"), "성명": idx("성명", "성 명", "성     명"),
        "대수": idx("대수"), "이월금": idx("이월금"),
    }
    month_cols: dict[int, int] = {}
    for i, h in enumerate(headers):
        mm = re.search(r"(\d{1,2})월\s*미수금", _clean(h).replace(" ", ""))
        if mm:
            month_cols[int(mm.group(1))] = i

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row_no + 1, max_col=80, values_only=True):
        vehicle = _clean(row[base["차량번호"]]) if base["차량번호"] is not None and base["차량번호"] < len(row) else ""
        name = _person_name(row[base["성명"]]) if base["성명"] is not None and base["성명"] < len(row) else ""
        if not _valid_vehicle(vehicle) or not name:
            continue
        last_month = 0
        current_amount = 0
        monthly_values: dict[str, int] = {}
        for month in range(1, 13):
            col = month_cols.get(month)
            val = row[col] if col is not None and col < len(row) else None
            money = _money(val)
            monthly_values[f"{month}월 미수금"] = money
            # 행별 마지막 입력 월: 0도 입력값이면 해당 월을 기준월로 본다.
            if _has_value(val):
                last_month = month
                current_amount = money
        out = {
            "지역": _clean(row[base["지역"]]) if base["지역"] is not None and base["지역"] < len(row) else "",
            "계정": _clean(row[base["계정"]]) if base["계정"] is not None and base["계정"] < len(row) else "",
            "비고": _clean(row[base["비고"]]) if base["비고"] is not None and base["비고"] < len(row) else "",
            "차량번호": vehicle,
            "성명": name,
            "대수": _clean(row[base["대수"]]) if base["대수"] is not None and base["대수"] < len(row) else "",
            "이월금": _money(row[base["이월금"]]) if base["이월금"] is not None and base["이월금"] < len(row) else 0,
            "마지막 미수 기준월": f"2026-{last_month:02d}" if last_month else "",
            "기준월": f"2026-{last_month:02d}" if last_month else "",
            "현재 미수금액": current_amount,
            "미수금액": current_amount,
            **monthly_values,
        }
        rows.append(out)
        if preview_limit and len(rows) >= preview_limit:
            return rows
    return rows


def _pick_col(keys: list[str], includes: list[str], excludes: list[str] | None = None) -> str | None:
    excludes = excludes or []
    for k in keys:
        nk = _norm_col(k)
        if any(x in nk for x in includes) and not any(x in nk for x in excludes):
            return k
    return None


def _iter_deposit_rows(file_bytes: bytes, preview_limit: int | None = None) -> list[dict[str, Any]]:
    """통장/카드 수납 엑셀을 수납 매칭용 공통 컬럼으로 정규화한다.

    원본 컬럼을 그대로 보여주는 화면은 매칭에 도움이 적으므로,
    거래일자/입금자명/입금액/거래내용/거래기록사항 중심으로 변환한다.
    카드매출 파일(승인일자, 정산일자, 카드사, 승인금액 등)도 수납 후보로 읽을 수 있게 한다.
    """
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, dtype=object)
    df = df.dropna(how="all").head(preview_limit or 20000)
    raw_rows = [{str(k): _json(v) for k, v in r.items()} for r in df.to_dict("records")]
    if not raw_rows:
        return []
    keys = list(raw_rows[0].keys())

    # 거래일자
    date_key = _pick_col(keys, ["거래일자", "입금일", "수납일", "승인일자", "정산일자", "날짜", "일자"])

    # 입금금액 우선, 출금금액/잔액은 절대 사용하지 않는다
    amt_key = None
    for candidate in ["입금금액", "입금액", "수납금액", "승인금액", "정산금액"]:
        found = _pick_col(keys, [candidate], excludes=["출금", "잔액", "수수료", "잔"])
        if found:
            amt_key = found
            break
    if not amt_key:
        # fallback: 금액 포함 컬럼 중 출금/잔액이 아닌 것
        amt_key = _pick_col(keys, ["금액"], excludes=["출금", "잔액", "수수료", "후"])

    # 거래기록사항(실제 입금자명) > 거래내용 순
    # 거래점(농협/신한 등 은행명)은 입금자명으로 쓰면 안 됨
    name_key = _pick_col(keys, ["거래기록사항", "기록사항", "입금자명", "입금자", "예금주"])
    fallback_name_key = _pick_col(keys, ["거래내용", "적요", "내용"], excludes=["거래점", "점"])

    memo_key = _pick_col(keys, ["거래내용", "적요", "내용", "메모"])
    location_key = _pick_col(keys, ["거래점", "가맹점"])
    balance_key = _pick_col(keys, ["거래 후 잔액", "잔액", "잔"])

    rows: list[dict[str, Any]] = []
    for r in raw_rows:
        amount = _money(r.get(amt_key)) if amt_key else 0
        if amount <= 0:
            # 입금금액 0 또는 없음 = 출금 행 → 수납매칭 제외
            continue

        # 입금자명: 거래기록사항 우선, 비어있으면 거래내용
        name = _clean(r.get(name_key)) if name_key else ""
        if not name and fallback_name_key:
            name = _clean(r.get(fallback_name_key))
        if not name:
            name = "입금자미상"

        needs_review = amount > 1_000_000

        rows.append({
            "거래일자": _clean(r.get(date_key)) if date_key else "",
            "입금자명": name,
            "입금액": amount,
            "출금액": 0,  # 출금행은 이미 제외됨
            "거래내용": _clean(r.get(memo_key)) if memo_key else "",
            "거래기록사항": _clean(r.get(name_key)) if name_key else "",
            "거래점": _clean(r.get(location_key)) if location_key else "",
            "원본구분": "카드매출" if any("카드" in _norm_col(k) or "승인" in _norm_col(k) for k in keys) else "통장거래",
            "원본메모": " / ".join(f"{k}:{_clean(v)}" for k, v in r.items() if _clean(v))[:900],
            "검토필요": "고액(100만원초과)" if needs_review else "",
        })
    return rows


def _member_match_maps(db: Session) -> tuple[dict[str, Member], dict[tuple[str, str], Member]]:
    by_vehicle: dict[str, Member] = {}
    by_name_last4: dict[tuple[str, str], Member] = {}
    for member in db.scalars(select(Member)).all():
        vn = _vehicle_norm(member.vehicle_no)
        if vn:
            by_vehicle[vn] = member
        last4 = _vehicle_last4(member.vehicle_no)
        nm = _person_name(member.name)
        if nm and last4:
            by_name_last4[(nm, last4)] = member
    return by_vehicle, by_name_last4


def _find_member_from_maps(by_vehicle: dict[str, Member], by_name_last4: dict[tuple[str, str], Member], name: str, vehicle_no: str) -> Member | None:
    vn = _vehicle_norm(vehicle_no)
    if vn and vn in by_vehicle:
        return by_vehicle[vn]
    key = (_person_name(name), _vehicle_last4(vehicle_no))
    if key[0] and key[1]:
        return by_name_last4.get(key)
    return None


@router.post("/preview")
async def preview_import(file_type: str = Form(...), file: UploadFile = File(...), db: Session = Depends(get_db)):
    data = await file.read()
    try:
        if file_type == "members":
            rows = _iter_license_rows(data, preview_limit=300)
            return {
                "ok": True,
                "filename": file.filename,
                "type": "members",
                "message": "전체면허자현황은 개인+택배 시트만 읽습니다. 업체세분/선진물류/업체/차량집계는 제외됩니다.",
                "total_rows": len(rows),
                "columns": STANDARD_MEMBER_COLUMNS,
                "raw_columns": ["개인 시트", "택배 시트"],
                "sample": rows,
            }
        if file_type == "arrears":
            rows = _iter_arrears_rows(data, preview_limit=300)
            by_vehicle, by_name_last4 = _member_match_maps(db)
            for r in rows:
                member = _find_member_from_maps(by_vehicle, by_name_last4, r.get("성명", ""), r.get("차량번호", ""))
                r["매칭상태"] = "매칭" if member else "미매칭"
                r["매칭된 회원명"] = member.name if member else ""
                r["매칭된 차량번호"] = member.vehicle_no if member else ""
            return {
                "ok": True,
                "filename": file.filename,
                "type": "arrears",
                "message": "2026미수금은 2026년회비내역 시트만 읽고, 월별 미수금을 합산하지 않습니다. 행별 마지막 입력 월의 미수금 1건만 현재잔액으로 반영합니다.",
                "total_rows": len(rows),
                "columns": ARREARS_PREVIEW_COLUMNS,
                "raw_columns": ["2026년회비내역"],
                "sample": rows,
            }
        if file_type == "deposits":
            rows = _iter_deposit_rows(data, preview_limit=300)
            return {
                "ok": True,
                "filename": file.filename,
                "type": "deposits",
                "message": "통장/카드 수납 엑셀은 거래일자·입금자명·입금액·거래내용으로 정규화됩니다. 저장 후 통장매칭 화면에서 자동매칭/후보확인/미매칭을 처리하세요.",
                "total_rows": len(rows),
                "columns": DEPOSIT_PREVIEW_COLUMNS,
                "raw_columns": list(rows[0].keys()) if rows else [],
                "sample": rows,
            }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"엑셀 미리보기 실패: {str(exc)[:500]}") from exc
    raise HTTPException(status_code=400, detail="file_type은 members / arrears / deposits 중 하나여야 합니다.")


@router.post("/reset")
def reset_misu_data(db: Session = Depends(get_db)):
    try:
        db.execute(delete(MemberHistory))
        db.execute(delete(Payment))
        db.execute(delete(ReceivableItem))
        db.execute(delete(Closure))
        db.execute(delete(Deposit))
        db.execute(delete(Pending))
        db.execute(delete(Member))
        db.commit()
        return {"ok": True, "message": "misu_* 업무 데이터 초기화 완료"}
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"초기화 실패: {str(exc)[:300]}") from exc


@router.post("/commit")
async def commit_import(file_type: str = Form(...), file: UploadFile = File(...), db: Session = Depends(get_db)):
    data = await file.read()
    if file_type == "members":
        rows = _iter_license_rows(data, preview_limit=None)
        inserted = updated = skipped = 0
        next_no = _current_max_member_no(db) + 1
        used_ids = set(db.scalars(select(Member.id)).all())
        used_mgmt: set[str] = {x for x in db.scalars(select(Member.mgmt_no)).all() if x}
        by_vehicle, by_name_last4 = _member_match_maps(db)
        for row in rows:
            vehicle = row["차량번호"]
            name = row["성명"]
            if not _valid_vehicle(vehicle) or not name:
                skipped += 1
                continue
            existing = _find_member_from_maps(by_vehicle, by_name_last4, name, vehicle)
            membership = row["가입여부"]
            item = charge_item(membership)
            cert_date = _parse_date(row.get("자격증명 발급일자"))
            join_date = _parse_date(row.get("가입일자"))
            start_ym = next_month_ym(join_date if membership == "협회가입" else cert_date)
            byear = _birth_year(row.get("주민등록번호", ""))
            age = (date.today().year - byear) if byear else None
            amount = monthly_charge(membership, age=age, birth_year=byear)
            memo = _ledger_note(row)
            if existing:
                m = existing
                raw_mgmt = _clean(row.get("관리번호"))
                if raw_mgmt and (raw_mgmt == m.mgmt_no or raw_mgmt not in used_mgmt):
                    if m.mgmt_no:
                        used_mgmt.discard(m.mgmt_no)
                    m.mgmt_no = raw_mgmt[:16]
                    used_mgmt.add(m.mgmt_no)
                m.reg_type = "양도양수" if str(m.mgmt_no).startswith("양") else "신규"
                m.name = name
                m.vehicle_no = vehicle
                m.phone = _clip(row.get("핸드폰") or row.get("전화번호"), 20)
                m.sigun = row.get("지역") or "미분류"
                m.region_raw = row.get("지역") or "미분류"
                m.member_type = row.get("회원구분") or "개인"
                m.membership = membership
                m.birth_year = byear
                m.cert_issue_date = cert_date
                m.assoc_join_date = join_date
                m.billing_start_ym = start_ym
                m.charge_item = item
                m.monthly_charge = amount
                m.status = m.status or "정상"
                m.cert_missing = cert_date is None
                cleaned_memo = _strip_generated_member_memo(m.memo)
                m.memo = _merge_member_notes(cleaned_memo, memo)
                updated += 1
            else:
                while _next_member_id_from_no(next_no) in used_ids:
                    next_no += 1
                mid = _next_member_id_from_no(next_no)
                raw_mgmt = _make_mgmt_no(row.get("관리번호", ""))
                mgmt = raw_mgmt
                if mgmt:
                    suffix = 2
                    while mgmt in used_mgmt:
                        tail = f"-{suffix}"
                        mgmt = f"{raw_mgmt[:16-len(tail)]}{tail}"
                        suffix += 1
                    used_mgmt.add(mgmt)
                m = Member(
                    id=mid,
                    mgmt_no=mgmt,
                    reg_type="양도양수" if (mgmt and mgmt.startswith("양")) else "신규",
                    name=name,
                    vehicle_no=vehicle,
                    phone=_clip(row.get("핸드폰") or row.get("전화번호"), 20),
                    sigun=row.get("지역") or "미분류",
                    region_raw=row.get("지역") or "미분류",
                    member_type=row.get("회원구분") or "개인",
                    membership=membership,
                    birth_year=byear,
                    cert_issue_date=cert_date,
                    assoc_join_date=join_date,
                    billing_start_ym=start_ym,
                    charge_item=item,
                    monthly_charge=amount,
                    status="정상",
                    is_disconnected=False,
                    cert_missing=cert_date is None,
                    memo=memo,
                )
                db.add(m)
                used_ids.add(mid)
                by_vehicle[_vehicle_norm(vehicle)] = m
                by_name_last4[(_person_name(name), _vehicle_last4(vehicle))] = m
                next_no += 1
                inserted += 1
            if (inserted + updated) % 500 == 0:
                db.commit()
        db.commit()
        return {"ok": True, "filename": file.filename, "type": "members", "inserted": inserted, "updated": updated, "skipped": skipped, "errors": []}

    if file_type == "arrears":
        rows = _iter_arrears_rows(data, preview_limit=None)
        inserted = updated = skipped = 0
        unmatched: list[str] = []
        by_vehicle, by_name_last4 = _member_match_maps(db)
        for row in rows:
            member = _find_member_from_maps(by_vehicle, by_name_last4, row["성명"], row["차량번호"])
            if not member:
                skipped += 1
                if len(unmatched) < 80:
                    unmatched.append(f"미매칭: {row['성명']} / {row['차량번호']} / {row.get('현재 미수금액', 0):,}원")
                continue

            # 2026미수금 파일의 비고는 회원 원장의 비고와 별개로 반드시 회원 메모에 합쳐 저장한다.
            # 통장매칭 후보/수동검색에서 이 메모를 보여주고, 메모 속 이름도 매칭 근거로 사용하기 위함.
            arrears_note = _clean(row.get("비고"))
            if arrears_note:
                member.memo = _append_unique_note(member.memo, "미수금 비고", arrears_note)
                if _is_contact_problem_note(arrears_note):
                    member.is_disconnected = True

            # 재업로드 시 해당 회원의 기존 미수 상세를 현재잔액 1건으로 교체한다.
            db.execute(delete(ReceivableItem).where(ReceivableItem.member_id == member.id))
            amt = _money(row.get("현재 미수금액") or row.get("미수금액"))
            ym = row.get("마지막 미수 기준월") or row.get("기준월") or "2026-현재"
            if amt <= 0:
                skipped += 1
                continue
            db.add(ReceivableItem(member_id=member.id, ym=ym, charge_item=member.charge_item, amount=amt, is_paid=False))
            inserted += 1
            if inserted % 500 == 0:
                db.commit()
        db.commit()
        return {"ok": True, "filename": file.filename, "type": "arrears", "inserted": inserted, "updated": updated, "skipped": skipped, "errors": unmatched}

    if file_type == "deposits":
        rows = _iter_deposit_rows(data, preview_limit=None)
        inserted = skipped = 0
        for r in rows:
            name = _clean(r.get("입금자명"))
            amt = _money(r.get("입금액"))
            if not name or amt <= 0:
                skipped += 1
                continue
            memo_parts = []
            if r.get("거래내용"):
                memo_parts.append(f"거래내용:{r.get('거래내용')}")
            if r.get("거래기록사항"):
                memo_parts.append(f"거래기록:{r.get('거래기록사항')}")
            if r.get("원본구분"):
                memo_parts.append(str(r.get("원본구분")))
            if r.get("원본메모"):
                memo_parts.append(str(r.get("원본메모"))[:300])
            db.add(Deposit(
                deposit_date=_parse_date(r.get("거래일자")) or date.today(),
                depositor_name=_clip(name, 40) or "입금자미상",
                amount=amt,
                memo=_clip(" / ".join(memo_parts), 60) or "엑셀 업로드",
                status="대기",
                is_excluded=False,
            ))
            inserted += 1
        db.commit()
        return {"ok": True, "filename": file.filename, "type": "deposits", "inserted": inserted, "updated": 0, "skipped": skipped, "errors": []}

    raise HTTPException(status_code=400, detail="file_type은 members / arrears / deposits 중 하나여야 합니다.")
